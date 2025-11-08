"""
12'de ki herşeyi yapacağız ancak bu işlemi Gemini ile değil, LLaMA ile yapacağız
yani
12 + LLaMA

on-prem bir sistem olacak
"""

from langchain_community.chat_models import ChatOllama
from langchain_core.prompts import PromptTemplate # Prompt yazabilmek için
from langchain_core.output_parsers import JsonOutputParser, PydanticOutputParser
from langchain_community.vectorstores import FAISS # verileri vektör veritabanına aktarmak için

# from langchain.chains import RetrievalQA de langchain.chains bulunamadığından bu 3 tanesini bunun için yazdık
from langchain_core.runnables import RunnablePassthrough
from langchain_core.output_parsers import StrOutputParser
from langchain_core.prompts import ChatPromptTemplate # retrieval: getirmek. Q:question. A:answer. sorunun cevabını getir

from langchain_community.embeddings import HuggingFaceEmbeddings # embedding yöntemimiz # embedding yapabilmek için
from sentence_transformers import SentenceTransformer

# txt'den çekmeyeceğimiz için: from langchain_community.document_loaders import TextLoader satırını silip bunu yazdım
from langchain_core.documents import Document  # LangChain'de metin parçalarını saklamak için kullanılan temel veri yapısı

import sqlite3
 
from pydantic import BaseModel, Field # çıktıyı yapılandırmak ve ayrıştırmak için

import os # .env dosyamızı proje içerisinde bulabilmek için

from openpyxl import Workbook, load_workbook # outputları excel'e kaydedebilmek için

from datetime import datetime # çıktıların zamanlarını tutabilemk için. ne zaman yazıldı bu diye

import warnings
warnings.filterwarnings("ignore") # uyarıları görmemek adına

# database'de ki verileri al
def dbden_veri_al(db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT baslik,icerik FROM urun_bilgileri")
    rows = cursor.fetchall()
    conn.close()
    
    return [Document(page_content=f"{baslik}:{icerik}") for baslik,icerik in rows] 

# geminimizi belirleyelim ve alalım
llm = ChatOllama(model="llama3.2:3b")

# embedding yapıyoruz
embedding_model = HuggingFaceEmbeddings(model_name = "sentence-transformers/all-MiniLM-L6-v2") 

# db'den text'i alıcaz(müşterinin sorularını cevaplmak adına metni modele verip rag yapmak için)
documents = dbden_veri_al("urun_bilgileri.db")
vector_store = FAISS.from_documents(documents, embedding_model) # embed edilmiş txt dosyasındaki kelimeleri vektör veritabınımıza FAISS yardımıyla kaydediyoruz
retriever = vector_store.as_retriever() 

# retriever chain
rag_chain = (
    {"context": retriever, "question": RunnablePassthrough()}
    | ChatPromptTemplate.from_template("Bağlam: {context}\n\nSoru: {question}")
    | llm
    | StrOutputParser()
)


# Output şemasını tanımla (şikayet kategorisi, yönlendirme, cevap)
class Response_Schema(BaseModel):
    kategori: str = Field(description="şikayetin kategorisi. örn 'ürün kategorisi', 'Teslimat Sorunu', 'Müşteri hizmetleri'")
    yonlendirme: str = Field(description="şikayetin yönlendirileceği birim. örn 'ürün departmanı', 'lojistik departmanı', 'Müşteri hizmetleri departmanı'")
    cevap: str = Field(description="kullanıcıya verilecek cevap. örn 'şikayetiniz alınmıştır', 'en kısa sürede dönüş yapılacaktır'")

"""
Cevabın şu formatta verilmesinden:
Kategori: Teslimat Sorunu
Yönlendirme: Müşteri Hizmetleri Departmanı
Cevap: Kargonuzun gecikmesi ve müşteri hizmetlerimize ulaşamamanızla ilgili şikayetiniz alınmıştır.
bu Response_Schema yapısı sorumlu. Bunu, bu yapıyor.
"""

# LLM'in çıktısını JSON formatına çevirmek için parser yarat
parser = PydanticOutputParser(pydantic_object=Response_Schema)

# LLM'e "şu formatta cevap ver" talimatını hazırla
formatted_instructions = parser.get_format_instructions()

# Prompt Şablonu
template = PromptTemplate(
    input_variables=["sikayet", "rag_bilgi"],
    partial_variables={ 
        "format_instructions":formatted_instructions
    }, 
    template="""
Sen bir müşteri hizmetleri asistanısın. Şikayeti analiz et ve SADECE JSON formatında cevap ver.

Bilgi Bankası:
{rag_bilgi}

KURALLAR:
- kategori: Kısa kategori adı (örn: "Ürün Sorunu", "Teslimat Sorunu")
- yonlendirme: İlgili departman (örn: "Ürün Departmanı", "Lojistik Departmanı")
- cevap: Müşteriye nazik ve profesyonel bir yanıt (özür + çözüm önerisi)

Şikayet: {sikayet}

{format_instructions}
"""
)

# kullancının mesajı
sikayet = input("Merhaba. Size nasıl yardımcı olabiliriz?: ")

rag_bilgi = rag_chain.invoke(sikayet)  # EKLE

prompt = template.format(sikayet=sikayet, rag_bilgi=rag_bilgi)  # rag_bilgi ekle

# llm'e prompt'u ver
response = llm.invoke(prompt)

result = parser.parse(response.content) # llm'in yanıtını ayrıştır

# Excel'e kaydet:
excel_file = "sikayet_raporu.xlsx"
dosya_var = os.path.exists(excel_file)

# Excel dosyasına yazma
if dosya_var:
    # Dosya varsa yükle
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    # Dosya yoksa yeni workbook oluştur ve başlıkları yaz
    wb = Workbook()
    ws = wb.active
    ws.append(["Zaman", "Şikayet", "Kategori", "Yönlendirme", "Cevap"])

# Verileri ekle
ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
           sikayet, result.kategori, result.yonlendirme, result.cevap])

# Kaydet
wb.save(excel_file)

print("\n--- Şikayet Analizi ---")
print("Kategori:", result.kategori)
print("Yönlendirme:", result.yonlendirme)
print("Cevap:", result.cevap)


"""
şu anda on-prem çalışan bir sistemimiz oldu. bilgisayarımızda kurulu llama modelimiz, dbmize bağlanıp rag yapıyor

örneğin dbmizde bunlar var:
    ("Kargo Süreci", "Kargom 3 gün içinde elime ulaştı, ancak gecikme yaşanabilir."),
    ("İade Politikası", "Ürünü 14 gün içinde koşulsuz iade edebilirsiniz."),
    ("Ödeme Seçenekleri", "Kredi kartı, havale ve kapıda ödeme seçenekleri mevcut."),
    ("Teknik Destek", "Ürün çalışmıyorsa destek hattını arayabilirsiniz."),
    ("İletişim", "Müşteri hizmetlerine hafta içi 09:00 - 18:00 arası ulaşabilirsiniz.")

şöyle sordum ve böyle cevap aldım:
Merhaba. Size nasıl yardımcı olabiliriz?: ödeme yöntemleriniz nelerdir?

--- Şikayet Analizi ---
Kategori: Müşteri Hizmetleri
Yönlendirme: Müşteri Hizmetleri Departmanı
Cevap: Ödeme yöntemlerimiz; Kredi kartı, Havale ve Kapıda ödeme seçeneği sunuyoruz. Lütfen seçiminizi belirtin.

YANİİİ KUSURSUZ.
"""