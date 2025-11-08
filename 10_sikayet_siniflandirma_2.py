"""
9'da yapmış olduğumuz kategorize etme projesini daha da geliştireceğiz. 

AI'ın yazdığı yanıtları, ve kullanıcı şikayetini bir excel dosyasına kaydedicez

Yani: 9.proje + raporlama olacak
"""

from langchain_google_genai import ChatGoogleGenerativeAI # Gemini'yi içeri aktardık
from langchain_core.prompts import PromptTemplate # Prompt yazabilmek için
from langchain_core.output_parsers import JsonOutputParser, PydanticOutputParser
from pydantic import BaseModel, Field # çıktıyı yapılandırmak ve ayrıştırmak için

from dotenv import load_dotenv # .env dosyamızı içeri aktarabilmek için

import os # .env dosyamızı proje içerisinde bulabilmek için

from openpyxl import Workbook, load_workbook # outputları excel'e kaydedebilmek için

from datetime import datetime # çıktıların zamanlarını tutabilemk için. ne zaman yazıldı bu diye

import warnings
warnings.filterwarnings("ignore") # uyarıları görmemek adına


# .env al
load_dotenv()
gemini_api_key = os.getenv("GEMINI_API_KEY")

# geminimizi belirleyelim ve alalım
llm = ChatGoogleGenerativeAI(model="models/gemini-2.5-flash", temperature=0.2, google_api_key=gemini_api_key)

# Output şemasını tanımla (şikayet kategorisi, yönlendirme, cevap)
class Response_Schema(BaseModel):
    kategori: str = Field(description="şikayetin kategorisi. örn 'ürün kategorisi', 'Teslimat Sorunu', 'Müşteri hizmetleri'")
    yonlendirme: str = Field(description="şikayetin yönlendirileceği birim. örn 'ürün departmanı', 'lojistik departmanı', 'Müşteri hizmetleri departmanı'")
    cevap: str = Field(description="kullanıcıya verilecek cevap. örn 'şikayetiniz alınmıştır', 'en kısa sürede dönüş yapılacaktır'")

"""
Cevabın şu formatta verilmesinden:

Kategori: Teslimat Sorunu
Yönlendirme: Müşteri Hizmetleri Departmanı
Cevap: Kargonuzun gecikmesi ve müşteri hizmetlerimize ulaşamamanızla ilgili şikayetiniz alınmıştır.

bu Response_Schema yapısı sorumlu. Bunu, bu yapıyor.
"""

# LLM'in çıktısını JSON formatına çevirmek için parser yarat
parser = PydanticOutputParser(pydantic_object=Response_Schema)

# LLM'e "şu formatta cevap ver" talimatını hazırla
formatted_instructions = parser.get_format_instructions()

# Prompt Şablonu
template = PromptTemplate(
    input_variables=["sikayet"], # kullanıcının mesajı
    partial_variables={ 
        "format_instructions":formatted_instructions # format talimatı
    }, 
    template="""
        Bir kullanıcıdan şikayeti aldın. 
        Bu şikayeti anlamlandır, uygun kategoriyi belirle, yönlendirme yap ve kullanıcıya kısa bir açıklama yaz

        {format_instructions}

        Şikayet: {sikayet}
        """
)

# kullancının mesajı
sikayet = input("Şikayetinizi Yazınız: ")

prompt = template.format(sikayet=sikayet) # prompt mesajını oluştur

# llm'e prompt'u ver
response = llm.invoke(prompt)

result = parser.parse(response.content) # llm'in yanıtını ayrıştır

# Excel'e kaydet:
excel_file = "sikayet_raporu.xlsx"
dosya_var = os.path.exists(excel_file)

# Excel dosyasına yazma
if dosya_var:
    # Dosya varsa yükle
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    # Dosya yoksa yeni workbook oluştur ve başlıkları yaz
    wb = Workbook()
    ws = wb.active
    ws.append(["Zaman", "Şikayet", "Kategori", "Yönlendirme", "Cevap"])

# Verileri ekle
ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
           sikayet, result.kategori, result.yonlendirme, result.cevap])

# Kaydet
wb.save(excel_file)

print("\n--- Şikayet Analizi ---")
print("Kategori:", result.kategori)
print("Yönlendirme:", result.yonlendirme)
print("Cevap:", result.cevap)